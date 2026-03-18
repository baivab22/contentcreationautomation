#!/usr/bin/env python3
"""
Instagram Profile Monitor — Core scraper module (instagrapi-based)
- Uses Instagram's Private Mobile API (fewer requests, less rate-limiting)
- One-time login with session persistence
- Fetches posts within a date range efficiently
- Downloads post images/videos
- Extracts text from images via OCR
- Saves everything to an Excel file with embedded media
"""

import base64
import json
import os
import sys
import time
import random
import re
import hashlib
import shutil
import threading
import requests as _requests
from datetime import datetime, timezone
from pathlib import Path

from instagrapi import Client
from instagrapi.types import Media
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
try:
    import pytesseract
    HAS_TESSERACT = True
except ImportError:
    HAS_TESSERACT = False

try:
    from google.oauth2 import service_account as google_service_account
    from googleapiclient.discovery import build as google_build
except ImportError:
    google_service_account = None
    google_build = None

# ─── Paths ────────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent
CONFIG_FILE = BASE_DIR / "config.json"
SESSION_DIR = BASE_DIR / "session"
SEEN_FILE = BASE_DIR / "seen_posts.json"

# ─── AI rewrite (OpenRouter) ────────────────────────────────────────────────
OPENROUTER_API_URL = "https://openrouter.ai/api/v1/chat/completions"
DEFAULT_OPENROUTER_MODEL = "deepseek/deepseek-r1"
DEFAULT_OPENROUTER_PROMPT = (
    "You are a senior SEO expert. Use simple and clear words. "
    "Generate a short title, a short description, and a one-sentence summary from the post caption."
)

_ai_cache: dict[str, dict] = {}
_ai_cache_lock = threading.Lock()
_ai_config_cache = None
_ai_config_mtime = None

# ─── Google Sheets export ───────────────────────────────────────────────────
GOOGLE_SHEETS_SCOPE = "https://www.googleapis.com/auth/spreadsheets"
DEFAULT_SHEETS_WORKSHEET = "Instagram Posts"

_sheets_lock = threading.Lock()
_sheets_config_cache = None
_sheets_config_mtime = None
_sheets_service_cache = None
_sheets_service_identity = None
_sheets_ready_targets: set[tuple[str, str]] = set()
_sheets_warned_keys: set[str] = set()
_config_warn_lock = threading.Lock()
_config_warned = False

# ─── Load config ──────────────────────────────────────────────────────────────
def _default_config() -> dict:
    return {
        "instagram_credentials": {"username": "", "password": ""},
        "instagram_allow_fresh_login": False,
        "profiles": [],
        "monitor_interval_seconds": 300,
        "download_media": False,
        "excel_file": "instagram_posts.xlsx",
        "media_folder": "downloaded_media",
        "ai": {
            "openrouter": {
                "api_key": "",
                "model": DEFAULT_OPENROUTER_MODEL,
                "prompt": DEFAULT_OPENROUTER_PROMPT,
                "timeout_seconds": 90,
                "temperature": 0.5,
            }
        },
        "publisher": {
            "sheets": {
                "enabled": True,
                "spreadsheet_id": "",
                "worksheet_name": DEFAULT_SHEETS_WORKSHEET,
                "credentials_file": "secrets/autoscraper-489906-6efe766866da.json",
            },
            "facebook": {
                "page_id": "",
                "access_token": "",
                "api_version": "v22.0",
            },
        },
    }


def _warn_config_once(message: str):
    global _config_warned
    with _config_warn_lock:
        if _config_warned:
            return
        _config_warned = True
    print(message)


def load_config():
    fallback = _default_config()

    try:
        with open(CONFIG_FILE, "r") as f:
            data = json.load(f)
    except Exception as e:
        _warn_config_once(
            f"[Config] Failed to load {CONFIG_FILE}: {e}. Using fallback defaults."
        )
        return fallback

    if not isinstance(data, dict):
        _warn_config_once(
            f"[Config] Invalid config root type ({type(data).__name__}). "
            "Using fallback defaults."
        )
        return fallback

    # Fill required keys so API routes can keep responding even if config is partial.
    data.setdefault("instagram_credentials", fallback["instagram_credentials"])
    data.setdefault("instagram_allow_fresh_login", fallback["instagram_allow_fresh_login"])
    data.setdefault("profiles", fallback["profiles"])
    data.setdefault("monitor_interval_seconds", fallback["monitor_interval_seconds"])
    data.setdefault("download_media", fallback["download_media"])
    data.setdefault("excel_file", fallback["excel_file"])
    data.setdefault("media_folder", fallback["media_folder"])

    ai_cfg = data.setdefault("ai", {})
    ai_openrouter = ai_cfg.setdefault("openrouter", {})
    for key, value in fallback["ai"]["openrouter"].items():
        ai_openrouter.setdefault(key, value)

    publisher = data.setdefault("publisher", {})

    sheets_cfg = publisher.setdefault("sheets", {})
    for key, value in fallback["publisher"]["sheets"].items():
        sheets_cfg.setdefault(key, value)

    facebook_cfg = publisher.setdefault("facebook", {})
    for key, value in fallback["publisher"]["facebook"].items():
        facebook_cfg.setdefault(key, value)

    return data


def load_service_account_info_from_env(raw_env_names, b64_env_names):
    ordered_env_names = []
    seen = set()

    for env_name in (*b64_env_names, *raw_env_names):
        if env_name in seen:
            continue
        seen.add(env_name)
        ordered_env_names.append(env_name)

    for env_name in ordered_env_names:
        raw_value = os.getenv(env_name, "")
        if not raw_value:
            continue

        # First, try parsing as plain JSON. This supports raw JSON being pasted
        # into either the *_JSON or *_JSON_B64 variable by mistake.
        try:
            data = json.loads(raw_value)
        except Exception:
            data = None

        if isinstance(data, dict):
            return data, env_name

        # Then try base64-decoding and parsing the decoded JSON. This supports
        # base64 content being pasted into either env-var name by mistake.
        compact = "".join(raw_value.strip().split())
        if compact.endswith("%"):
            compact = compact[:-1]

        try:
            decoded = base64.b64decode(compact, validate=False)
            data = json.loads(decoded.decode("utf-8"))
        except Exception:
            data = None

        if isinstance(data, dict):
            return data, env_name

    return None, None


def _as_bool(value, default=False):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "on"}
    return default


def _warn_sheets_once(key: str, message: str):
    with _sheets_lock:
        if key in _sheets_warned_keys:
            return
        _sheets_warned_keys.add(key)
    print(message)


def _normalize_text(value: str, max_len: int | None = None) -> str:
    text = (value or "").replace("\r\n", "\n").strip()
    if max_len is not None:
        return text[:max_len]
    return text


def _title_from_text(text: str) -> str:
    words = re.findall(r"[A-Za-z0-9]+", text or "")
    if not words:
        return "Social Update"
    return " ".join(words[:10]).strip().title()


def _description_from_text(text: str, max_len: int = 320) -> str:
    normalized = _normalize_text(text)
    if not normalized:
        return ""

    compact = re.sub(r"\s+", " ", normalized)
    sentences = re.split(r"(?<=[.!?])\s+", compact)
    summary = " ".join(sentences[:2]).strip() if sentences else compact
    return _normalize_text(summary, max_len)


def _one_sentence_summary(text: str, max_len: int = 220) -> str:
    normalized = _normalize_text(text)
    if not normalized:
        return ""

    compact = re.sub(r"\s+", " ", normalized)
    sentences = re.split(r"(?<=[.!?])\s+", compact)
    summary = (sentences[0] if sentences else compact).strip()
    if summary and summary[-1] not in ".!?":
        summary = summary.rstrip(" ,;:") + "."
    return _normalize_text(summary, max_len)


def _build_hashtags(*parts: str, limit: int = 10) -> str:
    source = " ".join([p for p in parts if p])
    tokens = re.findall(r"[A-Za-z][A-Za-z0-9_]{2,}", source)
    stop = {
        "the", "and", "with", "from", "that", "this", "your", "about", "into",
        "have", "just", "more", "what", "when", "where", "were", "will", "been",
        "for", "you", "are", "our", "new", "only", "post", "caption", "image",
    }

    picked = []
    for raw in tokens:
        token = raw.lower()
        if token in stop or len(token) < 4:
            continue
        hashtag = "#" + token.capitalize()
        if hashtag not in picked:
            picked.append(hashtag)
        if len(picked) >= limit:
            break

    defaults = ["#InstagramGrowth", "#FacebookGrowth", "#SocialMedia", "#ContentMarketing"]
    for tag in defaults:
        if len(picked) >= limit:
            break
        if tag not in picked:
            picked.append(tag)

    return " ".join(picked)


def _ensure_hashtags(caption: str, *parts: str) -> str:
    body = _normalize_text(caption, 2100)
    hashtag_count = len(re.findall(r"#[A-Za-z0-9_]+", body))
    if hashtag_count >= 5:
        return body

    hashtags = _build_hashtags(body, *parts, limit=10)
    if not hashtags:
        return body

    if not body:
        return hashtags
    if hashtags in body:
        return body

    return _normalize_text(f"{body}\n\n{hashtags}", 2200)


def _fallback_ai_rewrite(caption: str, ocr_text: str) -> dict:
    clean_caption = _normalize_text(caption, 1900)
    clean_ocr = _normalize_text(ocr_text, 1800)

    generated_caption = _ensure_hashtags(
        clean_caption or clean_ocr or "Fresh update",
        clean_caption,
        clean_ocr,
    )

    generated_description = _description_from_text(clean_caption or clean_ocr)
    if not generated_description:
        generated_description = "Simple update shared for social media audiences."

    generated_summary = _one_sentence_summary(clean_caption or clean_ocr or generated_description)
    if not generated_summary:
        generated_summary = _one_sentence_summary(generated_description)

    new_ocr = clean_ocr or _normalize_text(clean_caption, 600)
    title_source = new_ocr or clean_caption
    generated_title = _normalize_text(_title_from_text(title_source), 120)

    return {
        "generated_caption": _normalize_text(generated_caption, 2200),
        "generated_description": _normalize_text(generated_description, 420),
        "generated_summary": _normalize_text(generated_summary, 220),
        "generated_title": generated_title,
        "new_ocr_text": _normalize_text(new_ocr, 1800),
        "post_title": generated_title,
    }


def _extract_json_object(raw_text: str):
    text = (raw_text or "").strip()
    if not text:
        return None

    if text.startswith("```"):
        lines = text.splitlines()
        lines = [ln for ln in lines if not ln.startswith("```")]
        text = "\n".join(lines).strip()

    try:
        return json.loads(text)
    except json.JSONDecodeError:
        start = text.find("{")
        end = text.rfind("}")
        if start != -1 and end != -1 and end > start:
            try:
                return json.loads(text[start:end + 1])
            except json.JSONDecodeError:
                return None
    return None


def get_ai_rewrite_config():
    global _ai_config_cache, _ai_config_mtime

    config_mtime = None
    try:
        config_mtime = CONFIG_FILE.stat().st_mtime
    except FileNotFoundError:
        pass

    if _ai_config_cache is None or _ai_config_mtime != config_mtime:
        config = load_config()
        ai_cfg = config.get("ai", {}).get("openrouter", {})
        api_key = os.getenv("OPENROUTER_API_KEY") or ai_cfg.get("api_key", "")
        model = os.getenv("OPENROUTER_MODEL") or ai_cfg.get("model", DEFAULT_OPENROUTER_MODEL)
        prompt = ai_cfg.get("prompt") or DEFAULT_OPENROUTER_PROMPT

        _ai_config_cache = {
            "enabled": bool(api_key),
            "api_key": api_key,
            "model": model,
            "prompt": prompt,
            "timeout": int(ai_cfg.get("timeout_seconds", 90)),
            "temperature": float(ai_cfg.get("temperature", 0.5)),
        }
        _ai_config_mtime = config_mtime

    return _ai_config_cache


def get_google_sheets_config():
    global _sheets_config_cache, _sheets_config_mtime

    config_mtime = None
    try:
        config_mtime = CONFIG_FILE.stat().st_mtime
    except FileNotFoundError:
        pass

    if _sheets_config_cache is None or _sheets_config_mtime != config_mtime:
        config = load_config()
        publisher_cfg = config.get("publisher", {})
        sheets_cfg = publisher_cfg.get("sheets", {})

        env_enabled = os.getenv("GOOGLE_SHEETS_ENABLED")
        enabled = (
            _as_bool(env_enabled, default=False)
            if env_enabled is not None
            else _as_bool(sheets_cfg.get("enabled"), default=False)
        )

        _sheets_config_cache = {
            "enabled": enabled,
            "spreadsheet_id": (
                os.getenv("GOOGLE_SHEETS_SPREADSHEET_ID")
                or sheets_cfg.get("spreadsheet_id")
                or ""
            ).strip(),
            "worksheet_name": (
                os.getenv("GOOGLE_SHEETS_WORKSHEET")
                or sheets_cfg.get("worksheet_name")
                or DEFAULT_SHEETS_WORKSHEET
            ).strip() or DEFAULT_SHEETS_WORKSHEET,
            "credentials_file": (
                os.getenv("GOOGLE_SHEETS_SERVICE_ACCOUNT_FILE")
                or sheets_cfg.get("credentials_file")
                or ""
            ).strip(),
        }
        _sheets_config_mtime = config_mtime

    return _sheets_config_cache


def _resolve_credentials_file(path_value: str) -> Path:
    if not path_value:
        raise RuntimeError("Google Sheets credentials file is not configured")

    candidate = Path(path_value)
    if not candidate.is_absolute():
        candidate = BASE_DIR / candidate

    candidate = candidate.resolve()
    if not candidate.exists() or not candidate.is_file():
        raise RuntimeError(f"Google Sheets credentials file not found: {candidate}")

    return candidate


def _get_sheets_service(sheets_cfg: dict):
    global _sheets_service_cache, _sheets_service_identity

    if google_service_account is None or google_build is None:
        raise RuntimeError(
            "Google Sheets dependencies are missing. Install google-auth and "
            "google-api-python-client in the active environment."
        )

    service_account_info, env_name = load_service_account_info_from_env(
        raw_env_names=("GOOGLE_SHEETS_CREDS_JSON",),
        b64_env_names=("GOOGLE_SHEETS_CREDS_JSON_B64",),
    )

    credentials_path = None
    if service_account_info is not None:
        identity = f"env:{env_name}:{service_account_info.get('client_email', '')}"
    else:
        credentials_path = _resolve_credentials_file(sheets_cfg.get("credentials_file", ""))
        identity = str(credentials_path)

    with _sheets_lock:
        if _sheets_service_cache is None or _sheets_service_identity != identity:
            if service_account_info is not None:
                creds = google_service_account.Credentials.from_service_account_info(
                    service_account_info,
                    scopes=[GOOGLE_SHEETS_SCOPE],
                )
            else:
                creds = google_service_account.Credentials.from_service_account_file(
                    str(credentials_path),
                    scopes=[GOOGLE_SHEETS_SCOPE],
                )
            _sheets_service_cache = google_build(
                "sheets",
                "v4",
                credentials=creds,
                cache_discovery=False,
            )
            _sheets_service_identity = identity

    return _sheets_service_cache


def rewrite_with_ai(caption: str, ocr_text: str, profile: str = "") -> dict:
    baseline = _fallback_ai_rewrite(caption, ocr_text)
    cfg = get_ai_rewrite_config()

    cache_key = hashlib.sha1(f"{caption}||{ocr_text}".encode("utf-8")).hexdigest()
    with _ai_cache_lock:
        if cache_key in _ai_cache:
            return _ai_cache[cache_key]

    if not cfg.get("enabled"):
        with _ai_cache_lock:
            _ai_cache[cache_key] = baseline
        return baseline

    system_prompt = (
        "You are a senior SEO and social media ranking expert for Instagram and Facebook. "
        "Use simple, easy words that anyone can understand. Keep all facts truthful to the original. "
        "Use the original caption as the main source, and use OCR text only if caption is empty or unclear. "
        "Return strict JSON only with keys: generatedCaption, generatedTitle, generatedDescription, generatedSummary. "
        "generatedCaption must include 8-12 SEO-friendly hashtags at the end."
    )

    user_prompt = (
        f"Instruction: {cfg['prompt']}\n\n"
        f"Profile: {profile}\n"
        f"Original caption/description:\n{caption or '[empty]'}\n\n"
        f"Original OCR text:\n{ocr_text or '[empty]'}\n\n"
        "Output requirements:\n"
        "1) generatedCaption: rewritten SEO caption with 8-12 relevant hashtags.\n"
        "2) generatedTitle: very short, simple title, max 8 words.\n"
        "3) generatedDescription: short description in simple words, max 2 sentences.\n"
        "4) generatedSummary: one sentence summary in very simple words.\n"
        "Return strict JSON only."
    )

    payload = {
        "model": cfg["model"],
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "temperature": cfg["temperature"],
    }

    headers = {
        "Authorization": f"Bearer {cfg['api_key']}",
        "Content-Type": "application/json",
    }

    for attempt in range(2):
        try:
            resp = _requests.post(
                OPENROUTER_API_URL,
                headers=headers,
                json=payload,
                timeout=cfg["timeout"],
            )
            resp.raise_for_status()
            body = resp.json()
            content = (
                body.get("choices", [{}])[0]
                .get("message", {})
                .get("content", "")
            )

            parsed = _extract_json_object(content)
            if not isinstance(parsed, dict):
                raise ValueError("AI response is not valid JSON")

            generated_caption = parsed.get("generatedCaption") or parsed.get("generated_caption") or ""
            generated_title = (
                parsed.get("generatedTitle")
                or parsed.get("generated_title")
                or parsed.get("postTitle")
                or parsed.get("post_title")
                or ""
            )
            generated_description = (
                parsed.get("generatedDescription")
                or parsed.get("generated_description")
                or parsed.get("newOcrText")
                or parsed.get("new_ocr_text")
                or ""
            )
            generated_summary = (
                parsed.get("generatedSummary")
                or parsed.get("generated_summary")
                or parsed.get("simpleSummary")
                or parsed.get("simple_summary")
                or parsed.get("summary")
                or ""
            )
            new_ocr_text = parsed.get("newOcrText") or parsed.get("new_ocr_text") or ""
            if not new_ocr_text:
                new_ocr_text = generated_description

            normalized_caption = _normalize_text(generated_caption, 2200) or baseline["generated_caption"]
            normalized_title = _normalize_text(generated_title, 120) or baseline["generated_title"]
            normalized_description = _normalize_text(generated_description, 420) or baseline["generated_description"]
            normalized_summary = _normalize_text(generated_summary, 220) or baseline["generated_summary"]

            result = {
                "generated_caption": _ensure_hashtags(normalized_caption, caption, ocr_text, profile),
                "generated_description": normalized_description,
                "generated_summary": normalized_summary,
                "generated_title": normalized_title,
                "new_ocr_text": _normalize_text(new_ocr_text, 1800) or baseline["new_ocr_text"],
                "post_title": normalized_title,
            }

            with _ai_cache_lock:
                _ai_cache[cache_key] = result
            return result
        except Exception as e:
            if attempt == 0:
                time.sleep(1.2)
                continue
            print(f"  [AI] Rewrite fallback used: {e}")

    with _ai_cache_lock:
        _ai_cache[cache_key] = baseline
    return baseline

# ─── Session management ──────────────────────────────────────────────────────
def _is_challenge_error(e: Exception) -> bool:
    msg = str(e).lower()
    return (
        "challengeunknownstep" in type(e).__name__.lower()
        or "challenge" in msg
        or "checkpoint" in msg
    )


def _is_blacklist_error(e: Exception) -> bool:
    msg = str(e).lower()
    markers = (
        "blacklist",
        "ip address",
        "help you get back into your account",
    )
    return any(marker in msg for marker in markers)


def _is_transient_session_error(e: Exception) -> bool:
    msg = str(e).lower()
    error_name = type(e).__name__.lower()

    if any(marker in msg for marker in ("login_required", "badpassword", "challenge", "checkpoint")):
        return False

    transient_markers = (
        "timed out",
        "timeout",
        "connection",
        "remote end closed",
        "temporarily",
        "try again",
        "internal error",
        "503",
        "502",
        "504",
    )
    transient_error_names = (
        "connecttimeout",
        "readtimeout",
        "connectionerror",
        "httperror",
        "requestexception",
    )

    return any(marker in msg for marker in transient_markers) or any(
        marker in error_name for marker in transient_error_names
    )


def _looks_like_json_file(path: Path) -> bool:
    try:
        with open(path, "rb") as fh:
            head = fh.read(64)
    except Exception:
        return False

    if not head:
        return False

    stripped = head.lstrip()
    if not stripped:
        return False
    return stripped[:1] in {b"{", b"["}


def _persist_session_aliases(cl: Client, usernames: list[str]):
    for uname in usernames:
        cleaned = (uname or "").strip()
        if not cleaned:
            continue
        try:
            cl.dump_settings(str(SESSION_DIR / f"session-{cleaned}.json"))
        except Exception as e:
            print(f"[!] Could not persist session alias for @{cleaned}: {type(e).__name__}")


def _iter_session_candidates(username: str):
    preferred = [
        SESSION_DIR / f"session-{username}.json",
        SESSION_DIR / f"session-{username}",
    ]

    seen = set()
    for path in preferred:
        if path.exists() and path.is_file():
            if not _looks_like_json_file(path):
                continue
            resolved = str(path.resolve())
            if resolved not in seen:
                seen.add(resolved)
                yield path

    others = sorted(
        [p for p in SESSION_DIR.glob("session-*") if p.is_file()],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )

    for path in others:
        if not _looks_like_json_file(path):
            continue
        resolved = str(path.resolve())
        if resolved in seen:
            continue
        seen.add(resolved)
        yield path


def _try_load_session(session_path: Path):
    """Return (client, username, error) after attempting session validation."""
    cl = Client()
    cl.delay_range = [2, 5]
    try:
        cl.load_settings(str(session_path))
        info = cl.account_info()
        return cl, info.username, None
    except Exception as e:
        return None, None, e


def get_loader(config):
    """Create an instagrapi Client and login with session persistence.

    Strategy:
      1. Try the configured account's saved session file and validate it with
         account_info() without forcing a fresh login.
      2. If unavailable, try any other reusable session file in session/.
      3. Only then attempt fresh login with configured credentials.
      4. Challenge and blacklist errors return actionable guidance.
    """
    username = config["instagram_credentials"]["username"]
    password = config["instagram_credentials"]["password"]
    session_file = SESSION_DIR / f"session-{username}.json"
    SESSION_DIR.mkdir(exist_ok=True)

    allow_fresh_login = _as_bool(
        os.getenv("INSTAGRAM_ALLOW_FRESH_LOGIN"),
        default=_as_bool(config.get("instagram_allow_fresh_login"), default=False),
    )

    candidates = list(_iter_session_candidates(username))
    transient_candidates: list[Path] = []

    # ── 1. Try configured account session first, then other session files ───
    for idx, candidate in enumerate(candidates):
        if idx == 0:
            print(f"[*] Loading saved session for @{username}...")
        else:
            print(f"[*] Trying fallback session file: {candidate.name}")

        cl, active_username, error = _try_load_session(candidate)
        if cl is not None:
            _persist_session_aliases(cl, [username, active_username])
            if idx == 0:
                print(f"[+] Session valid for @{active_username}")
            else:
                print(
                    f"[+] Reused saved session from @{active_username} "
                    f"(configured login: @{username})"
                )
            return cl

        if error is not None and _is_challenge_error(error):
            candidate.unlink(missing_ok=True)
            raise RuntimeError(
                f"Instagram requires manual verification for @{username}. "
                "Log in via a browser or the official app and approve the "
                "security check, then try again."
            ) from error

        if error is not None and _is_transient_session_error(error):
            transient_candidates.append(candidate)
            print(
                f"[!] Session validation was transient for {candidate.name} "
                f"({type(error).__name__}). Will retry it without strict validation."
            )
            continue

        print(f"[!] Fallback session skipped ({type(error).__name__}).")

    # If session validation failed due transient network/API errors, keep using
    # the cached session instead of forcing a fresh login from a blocked IP.
    for candidate in transient_candidates:
        try:
            cl = Client()
            cl.delay_range = [2, 5]
            cl.load_settings(str(candidate))
            print(
                f"[+] Reused cached session from {candidate.name} "
                "without strict validation (transient API issue)."
            )
            _persist_session_aliases(cl, [username])
            return cl
        except Exception as e:
            print(f"[!] Deferred session load failed ({type(e).__name__}).")

    # ── 2. Fresh login (no reusable saved session) ──────────────────────────
    if not allow_fresh_login:
        candidate_names = [p.name for p in candidates]
        if candidate_names:
            raise RuntimeError(
                "No reusable Instagram session could be validated. "
                f"Checked: {', '.join(candidate_names)}. "
                "Fresh login is disabled to avoid IP-based blocks. "
                "Approve login in Instagram app/web and save a valid session file, "
                "or set INSTAGRAM_ALLOW_FRESH_LOGIN=true for one explicit fresh-login attempt."
            )
        raise RuntimeError(
            "No Instagram session file found in session/. "
            "Fresh login is disabled to avoid IP-based blocks. "
            "Add a valid session-<username>.json file, or set "
            "INSTAGRAM_ALLOW_FRESH_LOGIN=true for one explicit fresh-login attempt."
        )

    print(f"[*] Logging in fresh as @{username}...")
    cl = Client()
    cl.delay_range = [2, 5]
    try:
        cl.login(username, password)
        cl.dump_settings(str(session_file))
        print(f"[+] Login successful. Session saved.")
        return cl
    except Exception as e:
        if _is_challenge_error(e):
            raise RuntimeError(
                f"Instagram requires manual verification for @{username}. "
                "Log in via a browser or the official app and approve the "
                "security check, then try again."
            ) from e
        if _is_blacklist_error(e):
            raise RuntimeError(
                f"Instagram blocked login for @{username} from this network/IP. "
                "Use a different network (for example mobile hotspot), approve login in the "
                "Instagram app/web, or keep a reusable session file in session/ and retry."
            ) from e
        raise


# ─── Seen-posts tracking ─────────────────────────────────────────────────────
def load_seen_posts():
    if SEEN_FILE.exists():
        with open(SEEN_FILE, "r") as f:
            return set(json.load(f))
    return set()


def save_seen_posts(seen: set):
    with open(SEEN_FILE, "w") as f:
        json.dump(list(seen), f)


# ─── OCR ──────────────────────────────────────────────────────────────────────
def extract_text_from_image(image_path: str) -> str:
    """Run Tesseract OCR on an image and return extracted text."""
    if not HAS_TESSERACT:
        return ""
    try:
        img = PILImage.open(image_path)
        text = pytesseract.image_to_string(img)
        return text.strip()
    except Exception as e:
        return f"[OCR Error: {e}]"


# ─── Media download ──────────────────────────────────────────────────────────
def _download_url(url: str, filepath: Path):
    """Download a URL to a local file."""
    if filepath.exists():
        return
    resp = _requests.get(url, timeout=60)
    resp.raise_for_status()
    with open(filepath, "wb") as f:
        f.write(resp.content)


def is_carousel_media(media: Media) -> bool:
    """Best-effort carousel detection across instagrapi versions/payload shapes."""
    if getattr(media, "media_type", None) == 8:
        return True

    product_type = (getattr(media, "product_type", "") or "").lower()
    if "carousel" in product_type:
        return True

    resources = getattr(media, "resources", None) or []
    if len(resources) > 1:
        return True

    return False


def download_post_media(cl, media: Media, media_folder: Path, download_enabled: bool = True) -> list[dict]:
    """Download all media (images/videos) for a post. Returns list of {path, type}."""
    if not download_enabled:
        return []
    media_folder.mkdir(parents=True, exist_ok=True)
    media_files = []

    # Carousels are intentionally excluded from fetching.
    if is_carousel_media(media):
        return media_files

    code = media.code
    date_str = media.taken_at.strftime("%Y%m%d_%H%M%S")
    owner = media.user.username if media.user else "unknown"
    base_name = f"{owner}_{date_str}_{code}"

    if media.media_type == 2 and media.video_url:  # Video
        filename = f"{base_name}.mp4"
        filepath = media_folder / filename
        _download_url(str(media.video_url), filepath)
        media_files.append({"path": str(filepath), "type": "video"})
    else:  # Photo
        filename = f"{base_name}.jpg"
        filepath = media_folder / filename
        url = str(media.thumbnail_url)
        _download_url(url, filepath)
        media_files.append({"path": str(filepath), "type": "image"})

    return media_files


# ─── Excel management ────────────────────────────────────────────────────────
HEADERS = [
    "postlink",
    "original caption",
    "generated caption",
    "generated title",
    "generated description",
    "simple summary",
    "Profile",
    "Post Date (UTC)",
    "Shortcode",
    "Media Type",
    "Media File Path",
    "Embedded Image",
    "OCR Text from Image",
    "Scraped At",
]


def _sheet_name_literal(name: str) -> str:
    return "'" + (name or DEFAULT_SHEETS_WORKSHEET).replace("'", "''") + "'"


def _ensure_google_sheet_ready(service, spreadsheet_id: str, worksheet_name: str):
    target = (spreadsheet_id, worksheet_name)
    with _sheets_lock:
        if target in _sheets_ready_targets:
            return

    metadata = service.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        fields="sheets(properties(title))",
    ).execute()
    existing_titles = {
        (sheet.get("properties", {}) or {}).get("title", "")
        for sheet in metadata.get("sheets", [])
    }

    if worksheet_name not in existing_titles:
        service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": [{"addSheet": {"properties": {"title": worksheet_name}}}]},
        ).execute()

    header_last_col = get_column_letter(len(HEADERS))
    header_range = f"{_sheet_name_literal(worksheet_name)}!A1:{header_last_col}1"
    existing_header = service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id,
        range=header_range,
    ).execute().get("values", [])

    if not existing_header:
        service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=f"{_sheet_name_literal(worksheet_name)}!A1",
            valueInputOption="RAW",
            body={"values": [HEADERS]},
        ).execute()

    with _sheets_lock:
        _sheets_ready_targets.add(target)


def _to_sheet_row(row_values: dict) -> list:
    return [row_values.get(header, "") for header in HEADERS]


def append_rows_to_google_sheet(rows: list[dict]):
    if not rows:
        return

    sheets_cfg = get_google_sheets_config()
    if not sheets_cfg.get("enabled"):
        return

    spreadsheet_id = (sheets_cfg.get("spreadsheet_id") or "").strip()
    worksheet_name = (sheets_cfg.get("worksheet_name") or DEFAULT_SHEETS_WORKSHEET).strip()

    if not spreadsheet_id:
        _warn_sheets_once(
            "missing_spreadsheet_id",
            "[Sheets] Missing spreadsheet_id. Set GOOGLE_SHEETS_SPREADSHEET_ID "
            "or config.json -> publisher.sheets.spreadsheet_id",
        )
        return

    try:
        service = _get_sheets_service(sheets_cfg)
        _ensure_google_sheet_ready(service, spreadsheet_id, worksheet_name)
        service.spreadsheets().values().append(
            spreadsheetId=spreadsheet_id,
            range=f"{_sheet_name_literal(worksheet_name)}!A1",
            valueInputOption="RAW",
            insertDataOption="INSERT_ROWS",
            body={"values": [_to_sheet_row(item) for item in rows]},
        ).execute()
    except Exception as e:
        _warn_sheets_once(
            f"append_failed:{type(e).__name__}",
            f"[Sheets] Append failed: {e}",
        )

HEADER_WIDTHS = {
    "postlink": 42,
    "original caption": 55,
    "generated caption": 60,
    "generated title": 32,
    "generated description": 48,
    "simple summary": 44,
    "Profile": 18,
    "Post Date (UTC)": 22,
    "Shortcode": 16,
    "Media Type": 12,
    "Media File Path": 40,
    "Embedded Image": 25,
    "OCR Text from Image": 50,
    "Scraped At": 22,
}

LEGACY_HEADER_RENAMES = {
    "Post URL": "postlink",
    "Caption / Description": "original caption",
    "generate caption/description": "generated caption",
    "postTItle": "generated title",
    "newOcrText": "generated description",
    "generatedSummary": "simple summary",
    "summary": "simple summary",
}

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header_cell(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def _get_header_index_map(ws) -> dict[str, int]:
    index = {}
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col_idx).value
        if isinstance(value, str) and value.strip():
            index[value.strip()] = col_idx
    return index


def _ensure_headers(ws) -> bool:
    changed = False
    header_index = _get_header_index_map(ws)

    for legacy_name, canonical_name in LEGACY_HEADER_RENAMES.items():
        legacy_col = header_index.get(legacy_name)
        if not legacy_col or canonical_name in header_index:
            continue

        cell = ws.cell(row=1, column=legacy_col, value=canonical_name)
        _style_header_cell(cell)
        header_index.pop(legacy_name, None)
        header_index[canonical_name] = legacy_col
        changed = True

    for header in HEADERS:
        if header not in header_index:
            col_idx = ws.max_column + 1
            cell = ws.cell(row=1, column=col_idx, value=header)
            _style_header_cell(cell)
            header_index[header] = col_idx
            changed = True

    for header, width in HEADER_WIDTHS.items():
        col_idx = header_index.get(header)
        if not col_idx:
            continue
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    return changed


def get_or_create_workbook(excel_path: str):
    """Open existing workbook or create a new one with headers."""
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
        changed = _ensure_headers(ws)
        if changed:
            wb.save(excel_path)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Instagram Posts"
        for col_idx, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            _style_header_cell(cell)

        _ensure_headers(ws)

        wb.save(excel_path)

    return wb, ws


def make_thumbnail(image_path: str, max_size: int = 150) -> str | None:
    """Create a thumbnail for embedding in Excel."""
    try:
        thumb_dir = Path(image_path).parent / "thumbnails"
        thumb_dir.mkdir(exist_ok=True)
        thumb_path = thumb_dir / f"thumb_{Path(image_path).name}"

        if thumb_path.exists():
            return str(thumb_path)

        img = PILImage.open(image_path)
        img.thumbnail((max_size, max_size))

        if img.mode in ("RGBA", "P"):
            img = img.convert("RGB")

        thumb_path_jpg = thumb_path.with_suffix(".jpg")
        img.save(thumb_path_jpg, "JPEG", quality=85)
        return str(thumb_path_jpg)
    except Exception as e:
        print(f"  [!] Thumbnail error: {e}")
        return None


def append_post_to_excel(excel_path: str, row_data: dict, media_files: list[dict]):
    """Append one or more rows (one per media item) for a post to the Excel file."""
    wb, ws = get_or_create_workbook(excel_path)
    header_index = _get_header_index_map(ws)
    embedded_col_idx = header_index.get("Embedded Image", 7)
    rows_for_google_sheet = []

    # When media download is disabled, still export one row using post metadata.
    effective_media_files = media_files or [{"path": "", "type": "not-downloaded"}]

    for media in effective_media_files:
        next_row = ws.max_row + 1
        media_path = media.get("path", "")
        media_type = media.get("type", "not-downloaded")

        ocr_text = ""
        if media_type == "image":
            ocr_text = extract_text_from_image(media_path)

        ai_rewrite = rewrite_with_ai(
            caption=row_data.get("caption", ""),
            ocr_text=ocr_text,
            profile=row_data.get("profile", ""),
        )

        # Keep OCR column populated even when local OCR engine is unavailable.
        if media_type == "image":
            normalized_ocr = _normalize_text(ocr_text)
            if (not normalized_ocr) or normalized_ocr.startswith("[OCR Error"):
                ocr_text = ai_rewrite.get("new_ocr_text", "") or _normalize_text(
                    row_data.get("caption", ""),
                    800,
                )

        generated_title = ai_rewrite.get("generated_title") or ai_rewrite.get("post_title", "")
        generated_description = ai_rewrite.get("generated_description") or ai_rewrite.get("new_ocr_text", "")
        generated_summary = ai_rewrite.get("generated_summary") or _one_sentence_summary(
            row_data.get("caption", "") or generated_description
        )

        row_values = {
            "postlink": row_data.get("post_url", ""),
            "original caption": row_data.get("caption", ""),
            "generated caption": ai_rewrite.get("generated_caption", ""),
            "generated title": generated_title,
            "generated description": generated_description,
            "simple summary": generated_summary,
            "Profile": row_data.get("profile", ""),
            "Post Date (UTC)": row_data.get("post_date", ""),
            "Shortcode": row_data.get("shortcode", ""),
            "Media Type": media_type,
            "Media File Path": media_path,
            "Embedded Image": "",
            "OCR Text from Image": ocr_text,
            "Scraped At": row_data.get("scraped_at", ""),
        }
        rows_for_google_sheet.append(row_values)

        for header, val in row_values.items():
            col_idx = header_index.get(header)
            if not col_idx:
                continue
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER

        if media_type == "image" and media_path:
            thumb_path = make_thumbnail(media_path)
            if thumb_path:
                try:
                    img = XlImage(thumb_path)
                    img.width = 120
                    img.height = 120
                    cell_ref = f"{get_column_letter(embedded_col_idx)}{next_row}"
                    ws.add_image(img, cell_ref)
                    ws.row_dimensions[next_row].height = 100
                except Exception as e:
                    print(f"  [!] Could not embed image: {e}")
        elif media_type == "video" and media_path:
            ws.cell(
                row=next_row,
                column=embedded_col_idx,
                value=f"[Video: {Path(media_path).name}]",
            )

    wb.save(excel_path)
    print(f"  [Excel] Saved to {excel_path}")
    append_rows_to_google_sheet(rows_for_google_sheet)


# ─── Profile scraping ────────────────────────────────────────────────────────
def wait_with_jitter(base_seconds: float, jitter: float = 0.5):
    """Sleep for base_seconds +/- jitter fraction."""
    actual = base_seconds * (1 + random.uniform(-jitter, jitter))
    time.sleep(max(1, actual))


def is_login_session_error(error: Exception) -> bool:
    message = str(error).lower()
    markers = (
        "login_required",
        "challenge_required",
        "checkpoint_required",
        "please wait a few minutes",
        "badpassword",
        "consent_required",
    )
    return any(marker in message for marker in markers)


def scrape_profile_in_range(cl, username: str,
                            date_from: datetime, date_to: datetime,
                            excel_path: str, media_folder: Path,
                            progress_cb=None, download_enabled: bool = True):
    """
    Scrape posts from a profile within a date range using instagrapi.
    Uses user_medias_paginated — fetches ~20 posts per API call,
    far fewer requests than the web GraphQL API.
    Returns number of posts scraped.
    """
    count = 0
    def log(msg):
        print(msg)
        if progress_cb:
            progress_cb(msg)

    try:
        user_id = cl.user_id_from_username(username)
        user_info = cl.user_info(user_id)
        log(f"Checking @{username} ({user_info.media_count} total posts)...")

        end_cursor = ""
        first_page = True  # Instagram returns pinned posts (out-of-order) on the first page only
        while True:
            medias, end_cursor = cl.user_medias_paginated(user_id, 20, end_cursor=end_cursor)

            if not medias:
                break

            reached_older = False
            for media in medias:
                if is_carousel_media(media):
                    code = media.code or str(getattr(media, "pk", ""))
                    log(f"  Skipping carousel post: {code}")
                    continue

                post_date = media.taken_at
                if post_date.tzinfo is None:
                    post_date = post_date.replace(tzinfo=timezone.utc)

                if post_date > date_to:
                    continue

                if post_date < date_from:
                    if first_page:
                        # Pinned posts appear out-of-order at the top of the first page.
                        # Skip them without stopping pagination so we still reach recent posts.
                        continue
                    reached_older = True
                    break

                code = media.code
                log(f"  Found: {code} ({post_date.strftime('%Y-%m-%d %H:%M:%S')})")

                media_files = download_post_media(cl, media, media_folder, download_enabled=download_enabled)
                if download_enabled and not media_files:
                    log(f"  Skipping post with no downloadable media: {code}")
                    continue

                caption = media.caption_text or ""
                row_data = {
                    "profile": username,
                    "post_date": post_date.strftime("%Y-%m-%d %H:%M:%S"),
                    "shortcode": code,
                    "caption": caption[:2000],
                    "post_url": f"https://www.instagram.com/p/{code}/",
                    "scraped_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
                }
                append_post_to_excel(excel_path, row_data, media_files)
                count += 1
                log(f"  Saved post {count}: {code}")

            first_page = False

            if reached_older or not end_cursor:
                break

            wait_with_jitter(3)

    except Exception as e:
        log(f"Error for @{username}: {e}")
        if is_login_session_error(e):
            raise RuntimeError(str(e)) from e

    return count


def monitor_profile(cl, username: str, seen: set,
                    excel_path: str, media_folder: Path, max_posts: int = 5,
                    download_enabled: bool = True):
    """Check a profile for new posts not yet in seen set."""
    new_count = 0
    try:
        user_id = cl.user_id_from_username(username)
        user_info = cl.user_info(user_id)
        print(f"\n[*] Checking @{username} ({user_info.media_count} posts)...")

        medias, _ = cl.user_medias_paginated(user_id, max_posts * 2)

        for media in medias:
            post_id = str(media.pk)
            if post_id in seen:
                break

            if is_carousel_media(media):
                code = media.code or post_id
                print(f"  [*] Skipping carousel post: {code}")
                seen.add(post_id)
                continue

            code = media.code
            print(f"  [+] New post: {code} ({media.taken_at})")

            media_files = download_post_media(cl, media, media_folder, download_enabled=download_enabled)
            if download_enabled and not media_files:
                print(f"  [*] Skipping post with no downloadable media: {code}")
                seen.add(post_id)
                continue

            caption = media.caption_text or ""
            row_data = {
                "profile": username,
                "post_date": media.taken_at.strftime("%Y-%m-%d %H:%M:%S"),
                "shortcode": code,
                "caption": caption[:2000],
                "post_url": f"https://www.instagram.com/p/{code}/",
                "scraped_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
            }
            append_post_to_excel(excel_path, row_data, media_files)

            seen.add(post_id)
            new_count += 1

            if new_count >= max_posts:
                print(f"  [*] Reached max {max_posts} new posts for this cycle.")
                break

    except Exception as e:
        print(f"  [!] Error for @{username}: {e}")
        if is_login_session_error(e):
            raise RuntimeError(str(e)) from e

    return new_count


# ─── Main loop ────────────────────────────────────────────────────────────────
def main():
    config = load_config()
    excel_path = str(BASE_DIR / config.get("excel_file", "instagram_posts.xlsx"))
    media_folder = BASE_DIR / config.get("media_folder", "downloaded_media")
    interval = config.get("monitor_interval_seconds", 300)
    download_enabled = bool(config.get("download_media", True))
    media_folder.mkdir(exist_ok=True)

    profiles = [p["username"] for p in config["profiles"] if p.get("enabled") and p.get("username")]

    if not profiles:
        print("[!] No enabled profiles found in config.json")
        sys.exit(1)

    print(f"[*] Monitoring {len(profiles)} profiles: {', '.join(profiles)}")
    print(f"[*] Check interval: {interval}s")
    print(f"[*] Excel file: {excel_path}")
    print(f"[*] Media folder: {media_folder}")
    print()

    cl = get_loader(config)
    get_or_create_workbook(excel_path)

    seen = load_seen_posts()
    print(f"[*] Already seen {len(seen)} posts from previous runs.")

    cycle = 0
    while True:
        cycle += 1
        print(f"\n{'='*60}")
        print(f"  MONITOR CYCLE #{cycle}  —  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"{'='*60}")

        total_new = 0
        for i, username in enumerate(profiles):
            relogin_attempted = False
            while True:
                try:
                    new = monitor_profile(cl, username, seen, excel_path, media_folder,
                                         download_enabled=download_enabled)
                    total_new += new
                    save_seen_posts(seen)
                    break
                except Exception as e:
                    if (not relogin_attempted) and is_login_session_error(e):
                        relogin_attempted = True
                        print("[!] Session issue detected. Re-authenticating and retrying once...")
                        try:
                            config = load_config()
                            cl = get_loader(config)
                            print("[+] Re-authentication successful. Retrying now...")
                            continue
                        except Exception as relogin_error:
                            print(f"[!] Re-authentication failed: {relogin_error}")

                    print(f"[!] Error monitoring @{username}: {e}")
                    break

            delay = random.randint(15, 30)
            print(f"  [*] Waiting {delay}s before next profile...")
            time.sleep(delay)

        print(f"\n[*] Cycle #{cycle} complete. {total_new} new posts found.")
        print(f"[*] Total tracked posts: {len(seen)}")
        print(f"[*] Next check in {interval} seconds... (Ctrl+C to stop)")

        try:
            time.sleep(interval)
        except KeyboardInterrupt:
            print("\n[*] Stopping monitor. Goodbye!")
            save_seen_posts(seen)
            break


if __name__ == "__main__":
    main()
