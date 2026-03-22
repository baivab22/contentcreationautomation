#!/usr/bin/env python3
"""
Flask Web UI for Instagram Profile Monitor
Run: python app.py
Open: http://localhost:5001
"""

import json
import os
import threading
import time
from datetime import datetime, timezone, timedelta
from pathlib import Path
from collections import deque

from flask import Flask, render_template, request, jsonify, send_file

from scraper import (
    BASE_DIR, CONFIG_FILE, load_config, get_loader,
    get_or_create_workbook, scrape_profile_in_range,
    load_seen_posts, save_seen_posts, monitor_profile,
    wait_with_jitter,
)

app = Flask(__name__, template_folder="templates", static_folder="static")

# ─── Global state ─────────────────────────────────────────────────────────────
_loader = None
_loader_identity = None
_loader_lock = threading.Lock()
_config = None
_config_mtime = None

# Job tracking
_jobs = {}  # job_id -> {status, logs, progress, total, done}
_job_counter = 0
_job_lock = threading.Lock()

FETCH_EXPORTS_DIR = BASE_DIR / "fetch_exports"
INSTAGRAM_MAX_CAPTION_LENGTH = 2200

_scrape_runtime_lock = threading.Lock()
_monitor_thread = None
_monitor_lock = threading.Lock()
_monitor_stop_event = threading.Event()


def get_config():
    global _config, _config_mtime
    config_mtime = None
    try:
        config_mtime = CONFIG_FILE.stat().st_mtime
    except FileNotFoundError:
        pass

    if _config is None or _config_mtime != config_mtime:
        _config = load_config()
        _config_mtime = config_mtime
    return _config


def _instagram_identity(config: dict):
    creds = config.get("instagram_credentials", {})
    return (
        (creds.get("username") or "").strip(),
        creds.get("password") or "",
    )


def get_instaloader():
    """Get or create a shared Instaloader instance (thread-safe)."""
    global _loader, _loader_identity
    config = get_config()
    identity = _instagram_identity(config)

    with _loader_lock:
        if _loader is None or _loader_identity != identity:
            _loader = get_loader(config)
            _loader_identity = identity
        return _loader


def reset_instagram_loader():
    global _loader, _loader_identity
    with _loader_lock:
        _loader = None
        _loader_identity = None


def new_job_id():
    global _job_counter
    with _job_lock:
        _job_counter += 1
        return f"job-{_job_counter}"


def get_enabled_profile_usernames():
    config = get_config()
    return [
        p["username"] for p in config.get("profiles", [])
        if p.get("username") and p.get("enabled")
    ]


def get_excel_reference(excel_path: Path) -> str:
    try:
        return excel_path.resolve().relative_to(BASE_DIR.resolve()).as_posix()
    except ValueError:
        return excel_path.name


def resolve_excel_path(file_ref: str | None) -> Path:
    if not file_ref:
        config = get_config()
        return (BASE_DIR / config.get("excel_file", "instagram_posts.xlsx")).resolve()

    candidate = Path(file_ref.strip())
    if candidate.is_absolute():
        raise ValueError("Absolute file paths are not allowed")

    resolved = (BASE_DIR / candidate).resolve()
    if not resolved.is_relative_to(BASE_DIR.resolve()):
        raise ValueError("Invalid Excel file path")
    if resolved.suffix.lower() != ".xlsx":
        raise ValueError("Only .xlsx files are supported")
    return resolved


def build_recent_fetch_excel_path(hours: int) -> Path:
    FETCH_EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    return FETCH_EXPORTS_DIR / f"recent_{hours}h_{stamp}.xlsx"


def create_scrape_job(selected_profiles, date_from: datetime, date_to: datetime,
                      excel_path: Path, job_label: str, fresh_output: bool = False):
    excel_path = excel_path.resolve()
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    excel_ref = get_excel_reference(excel_path)

    job_id = new_job_id()
    _jobs[job_id] = {
        "status": "running",
        "logs": deque(maxlen=500),
        "total_profiles": len(selected_profiles),
        "profiles_done": 0,
        "posts_found": 0,
        "current_profile": "",
        "excel_file": excel_ref,
        "job_label": job_label,
    }

    def run_job():
        config = get_config()
        media_folder = BASE_DIR / config.get("media_folder", "downloaded_media")
        media_folder.mkdir(exist_ok=True)
        # UI scraping should always download media for matched posts.
        download_enabled = True

        if fresh_output and excel_path.exists():
            try:
                excel_path.unlink()
                job = _jobs[job_id]
                job["logs"].append("Previous export file cleared. Starting fresh output.")
            except Exception as e:
                job = _jobs[job_id]
                job["logs"].append(f"Could not clear old export file: {e}")

        get_or_create_workbook(str(excel_path))

        job = _jobs[job_id]

        try:
            with _scrape_runtime_lock:
                get_instaloader()
        except Exception as e:
            job["logs"].append(f"Login failed: {e}")
            job["status"] = "error"
            return

        for i, username in enumerate(selected_profiles):
            job["current_profile"] = username
            job["logs"].append(f"--- Starting @{username} ({i+1}/{len(selected_profiles)}) ---")

            def progress_cb(msg):
                job["logs"].append(msg)

            relogin_attempted = False
            while True:
                try:
                    with _scrape_runtime_lock:
                        L = get_instaloader()
                        count = scrape_profile_in_range(
                            L, username, date_from, date_to,
                            str(excel_path), media_folder, progress_cb=progress_cb,
                            download_enabled=download_enabled
                        )
                    job["posts_found"] += count
                    job["logs"].append(f"@{username}: {count} posts scraped")
                    break
                except Exception as e:
                    if is_challenge_error(e):
                        job["logs"].append(
                            f"Instagram requires manual verification. "
                            "Please open the Instagram app or instagram.com, approve the "
                            "security check for the scraper account, then try again."
                        )
                        job["status"] = "error"
                        return
                    if (not relogin_attempted) and is_retryable_instagram_error(e):
                        relogin_attempted = True
                        job["logs"].append(
                            "Session issue detected. Re-authenticating and retrying this profile once..."
                        )
                        try:
                            with _scrape_runtime_lock:
                                reset_instagram_loader()
                                get_instaloader()
                            job["logs"].append("Re-authentication successful. Retrying now...")
                            continue
                        except Exception as relogin_error:
                            if is_challenge_error(relogin_error):
                                job["logs"].append(
                                    "Instagram requires manual verification. "
                                    "Please open the Instagram app or instagram.com, approve the "
                                    "security check for the scraper account, then try again."
                                )
                                job["status"] = "error"
                                return
                            job["logs"].append(f"Re-authentication failed: {relogin_error}")

                    job["logs"].append(f"Error for @{username}: {e}")
                    break

            job["profiles_done"] = i + 1

            if i < len(selected_profiles) - 1:
                import random
                delay = random.randint(15, 30)
                job["logs"].append(f"Waiting {delay}s before next profile...")
                time.sleep(delay)

        job["status"] = "completed"
        job["current_profile"] = ""
        job["logs"].append(f"=== Done! {job['posts_found']} total posts scraped ===")
        job["logs"].append(f"Output Excel: {excel_ref}")

    thread = threading.Thread(target=run_job, daemon=True)
    thread.start()

    return job_id, excel_ref


def utc_now_text():
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")


def as_bool(value, default=False):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "on"}
    return default


def should_run_background_monitor() -> bool:
    """Background monitor enabled by default. Only disable if explicitly set to false."""
    env_value = os.getenv("ENABLE_BACKGROUND_MONITOR", "true")
    result = as_bool(env_value, default=True)
    return result


def run_background_monitor_loop():
    config = get_config()
    excel_path = str(BASE_DIR / config.get("excel_file", "instagram_posts.xlsx"))
    media_folder = BASE_DIR / config.get("media_folder", "downloaded_media")
    media_folder.mkdir(exist_ok=True)
    download_enabled = bool(config.get("download_media", True))
    get_or_create_workbook(excel_path)

    seen = load_seen_posts()
    cycle = 0

    print("[Monitor] Background monitoring enabled.")
    print(f"[Monitor] Tracking {len(get_enabled_profile_usernames())} enabled profiles.")

    while not _monitor_stop_event.is_set():
        cycle += 1
        profiles = get_enabled_profile_usernames()

        total_new = 0
        if not profiles:
            print("[Monitor] No enabled profiles configured. Waiting for next cycle.")

        for index, username in enumerate(profiles):
            if _monitor_stop_event.is_set():
                break

            relogin_attempted = False
            while True:
                try:
                    with _scrape_runtime_lock:
                        L = get_instaloader()
                        new_count = monitor_profile(
                            L,
                            username,
                            seen,
                            excel_path,
                            media_folder,
                            download_enabled=download_enabled,
                        )

                    total_new += new_count
                    save_seen_posts(seen)
                    break
                except Exception as e:
                    if is_challenge_error(e):
                        print(
                            "[Monitor] Instagram requires manual verification. "
                            "Approve login in the app/web, then restart monitoring."
                        )
                        break

                    if (not relogin_attempted) and is_retryable_instagram_error(e):
                        relogin_attempted = True
                        print(f"[Monitor] Session issue for @{username}. Re-authenticating once...")
                        try:
                            with _scrape_runtime_lock:
                                reset_instagram_loader()
                                get_instaloader()
                            continue
                        except Exception as relogin_error:
                            print(f"[Monitor] Re-authentication failed for @{username}: {relogin_error}")

                    print(f"[Monitor] Error for @{username}: {e}")
                    break

            if index < len(profiles) - 1 and not _monitor_stop_event.is_set():
                wait_with_jitter(12, jitter=0.25)

        save_seen_posts(seen)
        interval = int(get_config().get("monitor_interval_seconds", 300))
        print(
            f"[Monitor] Cycle #{cycle} complete. New posts: {total_new}. "
            f"Next cycle in {interval}s."
        )

        slept = 0
        while slept < interval and not _monitor_stop_event.is_set():
            time.sleep(1)
            slept += 1


def start_background_monitor():
    global _monitor_thread

    if not should_run_background_monitor():
        print("[Monitor] Disabled via ENABLE_BACKGROUND_MONITOR.")
        return

    with _monitor_lock:
        if _monitor_thread is not None and _monitor_thread.is_alive():
            return

        _monitor_stop_event.clear()
        _monitor_thread = threading.Thread(
            target=run_background_monitor_loop,
            name="instagram-background-monitor",
            daemon=True,
        )
        _monitor_thread.start()


def normalize_instagram_caption(caption: str) -> str:
    cleaned = (caption or "").replace("\r\n", "\n").strip()
    return cleaned[:INSTAGRAM_MAX_CAPTION_LENGTH]


def is_challenge_error(error: Exception) -> bool:
    """Errors that require manual human verification — must NOT be auto-retried."""
    msg = str(error).lower()
    name = type(error).__name__.lower()
    return (
        "challengeunknownstep" in name
        or "challenge" in msg
        or "checkpoint" in msg
        or "manual verification" in msg
    )


def is_retryable_instagram_error(error: Exception) -> bool:
    if is_challenge_error(error):
        return False  # challenge errors need manual action, never auto-retry
    message = str(error).lower()
    markers = (
        "login_required",
        "badpassword",
        "please wait a few minutes",
        "feedback_required",
        "sentry_block",
        "consent_required",
        "timed out",
        "connection reset",
        "connection aborted",
        "temporary failure",
    )
    return any(marker in message for marker in markers)


# ─── Routes ───────────────────────────────────────────────────────────────────

@app.route("/healthz")
def healthz():
    return jsonify({"status": "ok"}), 200

@app.route("/api/publisher/items")
def publisher_items():
    """Safe endpoint - returns empty items. Google Drive not used."""
    return jsonify({
        "count": 0,
        "items": [],
        "message": "Google Drive integration disabled. Using Google Sheets for data export."
    }), 200

@app.route("/api/monitor/status")
def monitor_status():
    """Get background monitor status."""
    global _monitor_thread
    is_running = _monitor_thread is not None and _monitor_thread.is_alive()
    enabled = should_run_background_monitor()
    profiles = get_enabled_profile_usernames()
    return jsonify({
        "enabled": enabled,
        "running": is_running,
        "profiles_count": len(profiles),
        "profiles": profiles,
        "interval_seconds": get_config().get("monitor_interval_seconds", 300),
        "message": "Background monitor is active and running continuously"
    }), 200

@app.route("/")
def index():
    config = get_config()
    profiles = [p for p in config["profiles"] if p.get("username") and p.get("enabled")]
    excel_path = str(BASE_DIR / config.get("excel_file", "instagram_posts.xlsx"))
    excel_exists = os.path.exists(excel_path)

    # Count rows in excel
    row_count = 0
    if excel_exists:
        try:
            from openpyxl import load_workbook as _lw
            wb = _lw(excel_path)
            ws = wb.active
            row_count = max(0, ws.max_row - 1)  # minus header
        except Exception:
            pass

    return render_template("index.html",
                           profiles=profiles,
                           excel_exists=excel_exists,
                           row_count=row_count)


@app.route("/api/scrape", methods=["POST"])
def api_scrape():
    """Start a scraping job for selected profiles within a time range."""
    data = request.get_json(silent=True) or {}
    if not data:
        return jsonify({"error": "No JSON data provided"}), 400

    selected_profiles = data.get("profiles", [])
    date_from_str = data.get("date_from", "")
    date_to_str = data.get("date_to", "")

    if not selected_profiles:
        return jsonify({"error": "No profiles selected"}), 400
    if not date_from_str or not date_to_str:
        return jsonify({"error": "Both date_from and date_to are required"}), 400

    try:
        date_from = datetime.strptime(date_from_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        date_to = datetime.strptime(date_to_str, "%Y-%m-%d").replace(
            hour=23, minute=59, second=59, tzinfo=timezone.utc)
    except ValueError:
        return jsonify({"error": "Invalid date format. Use YYYY-MM-DD"}), 400

    if date_from > date_to:
        return jsonify({"error": "date_from must be before date_to"}), 400

    excel_path = resolve_excel_path(None)
    job_id, excel_ref = create_scrape_job(
        selected_profiles,
        date_from,
        date_to,
        excel_path=excel_path,
        job_label="custom-range",
        fresh_output=True,
    )

    return jsonify({"job_id": job_id, "status": "started", "excel_file": excel_ref})


@app.route("/api/fetch/recent", methods=["POST"])
def api_fetch_recent():
    """Fetch recent posts for all listed (enabled) accounts into a fresh Excel file."""
    data = request.get_json(silent=True) or {}

    try:
        hours = int(data.get("hours", 1))
    except (TypeError, ValueError):
        return jsonify({"error": "hours must be an integer"}), 400

    if hours < 1 or hours > 168:
        return jsonify({"error": "hours must be between 1 and 168"}), 400

    selected_profiles = get_enabled_profile_usernames()
    if not selected_profiles:
        return jsonify({"error": "No enabled profiles available in config"}), 400

    date_to = datetime.now(timezone.utc)
    date_from = date_to - timedelta(hours=hours)
    excel_path = build_recent_fetch_excel_path(hours)

    job_id, excel_ref = create_scrape_job(
        selected_profiles,
        date_from,
        date_to,
        excel_path=excel_path,
        job_label=f"recent-{hours}h-all",
        fresh_output=True,
    )

    return jsonify({
        "job_id": job_id,
        "status": "started",
        "hours": hours,
        "profiles": len(selected_profiles),
        "excel_file": excel_ref,
    })


@app.route("/api/job/<job_id>")
def api_job_status(job_id):
    """Get status of a scraping job."""
    job = _jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404

    return jsonify({
        "status": job["status"],
        "logs": list(job["logs"]),
        "total_profiles": job["total_profiles"],
        "profiles_done": job["profiles_done"],
        "posts_found": job["posts_found"],
        "current_profile": job["current_profile"],
        "excel_file": job.get("excel_file", ""),
        "job_label": job.get("job_label", ""),
    })


@app.route("/api/profiles")
def api_profiles():
    """Get list of configured profiles."""
    config = get_config()
    profiles = [p for p in config["profiles"] if p.get("username") and p.get("enabled")]
    return jsonify(profiles)


@app.route("/api/excel-info")
def api_excel_info():
    """Get info about current Excel file."""
    file_ref = (request.args.get("file") or "").strip()
    try:
        excel_path = resolve_excel_path(file_ref or None)
    except ValueError as e:
        return jsonify({"exists": False, "rows": 0, "size_kb": 0, "error": str(e)}), 400

    excel_ref = get_excel_reference(excel_path)
    if not os.path.exists(excel_path):
        return jsonify({"exists": False, "rows": 0, "size_kb": 0, "file": excel_ref})

    try:
        from openpyxl import load_workbook as _lw
        wb = _lw(str(excel_path))
        ws = wb.active
        rows = max(0, ws.max_row - 1)
    except Exception:
        rows = 0

    size_kb = round(os.path.getsize(str(excel_path)) / 1024, 1)
    return jsonify({"exists": True, "rows": rows, "size_kb": size_kb, "file": excel_ref})


@app.route("/download")
def download_excel():
    """Download the Excel file."""
    file_ref = (request.args.get("file") or "").strip()
    try:
        excel_path = resolve_excel_path(file_ref or None)
    except ValueError as e:
        return str(e), 400

    if not os.path.exists(excel_path):
        return "No Excel file yet. Run a scrape first.", 404

    download_name = excel_path.name if file_ref else "instagram_posts.xlsx"
    return send_file(str(excel_path), as_attachment=True,
                     download_name=download_name)


@app.route("/api/excel-data")
def api_excel_data():
    """Return Excel rows as JSON for the data table preview."""
    file_ref = (request.args.get("file") or "").strip()
    try:
        excel_path = resolve_excel_path(file_ref or None)
    except ValueError as e:
        return jsonify({"rows": [], "total": 0, "error": str(e)}), 400

    excel_ref = get_excel_reference(excel_path)
    if not os.path.exists(excel_path):
        return jsonify({"rows": [], "total": 0, "file": excel_ref})

    page = max(1, request.args.get("page", 1, type=int))
    per_page = min(100, max(1, request.args.get("per_page", 25, type=int)))

    try:
        from openpyxl import load_workbook as _lw
        wb = _lw(str(excel_path))
        ws = wb.active
        total_rows = max(0, ws.max_row - 1)

        headers = [cell.value for cell in ws[1]]
        # Read rows in reverse (newest first)
        start_row = max(2, ws.max_row - (page * per_page) + 1)
        end_row = min(ws.max_row, ws.max_row - ((page - 1) * per_page))

        rows = []
        for r in range(end_row, start_row - 1, -1):
            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                val = ws.cell(row=r, column=col_idx).value
                if header and header != "Embedded Image":
                    row_data[header] = str(val) if val is not None else ""
            rows.append(row_data)

        return jsonify({
            "rows": rows,
            "total": total_rows,
            "page": page,
            "per_page": per_page,
            "file": excel_ref,
        })
    except Exception as e:
        return jsonify({"rows": [], "total": 0, "error": str(e), "file": excel_ref})


@app.route("/media/<path:filename>")
def serve_media(filename):
    """Serve downloaded media files."""
    config = get_config()
    media_folder = BASE_DIR / config.get("media_folder", "downloaded_media")
    filepath = media_folder / filename
    # Prevent path traversal
    if not filepath.resolve().is_relative_to(media_folder.resolve()):
        return "Forbidden", 403
    if not filepath.exists():
        return "File not found", 404
    return send_file(str(filepath))


if __name__ == "__main__":
    print("=" * 50)
    print("  Instagram Monitor — Web UI")
    print("  Open http://localhost:5001")
    print("=" * 50)
    
    # Verify Google Sheets is enabled
    config = get_config()
    sheets_enabled = config.get("publisher", {}).get("sheets", {}).get("enabled", False)
    sheet_id = config.get("publisher", {}).get("sheets", {}).get("spreadsheet_id", "").strip()
    if sheets_enabled and sheet_id:
        print(f"✓ Google Sheets integration: ENABLED")
        print(f"  Spreadsheet ID: {sheet_id[:20]}...")
        print(f"  Target profiles: {len(get_enabled_profile_usernames())}")
    else:
        print("⚠ Google Sheets integration: DISABLED (no spreadsheet_id configured)")
    
    # Start background monitor
    print("\nStarting background monitor...")
    start_background_monitor()
    
    app_port = int(os.getenv("PORT", "5001"))
    print(f"\n✓ Server starting on http://0.0.0.0:{app_port}")
    print(f"✓ Health check: http://localhost:{app_port}/healthz")
    print(f"✓ Monitor status: http://localhost:{app_port}/api/monitor/status")
    print(f"✓ Profiles API: http://localhost:{app_port}/api/profiles")
    print("\nBackground monitoring is ACTIVE - data will be collected 24/7\n")
    
    app.run(host="0.0.0.0", port=app_port, debug=False)
